"""Хранение лидов: CSV/XLSX на выходе, SQLite для дедупликации по хэшу."""

import csv
import hashlib
import logging
import os
import sqlite3
from datetime import datetime, timezone

logger = logging.getLogger("storage")

CSV_COLUMNS = [
    "Дата",
    "Источник",
    "Название группы",
    "Автор",
    "Текст сообщения",
    "Ссылка",
    "Совпавшие ключевые слова",
    "Уверенность",
    "Статус",
]


def init_db(db_path="leads.db"):
    conn = sqlite3.connect(db_path)
    conn.execute(
        "CREATE TABLE IF NOT EXISTS seen_leads ("
        "hash TEXT PRIMARY KEY, "
        "created_at TEXT NOT NULL"
        ")"
    )
    conn.commit()
    return conn


def lead_hash(author, text):
    payload = f"{author or ''}|{text or ''}".strip().lower()
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def is_duplicate(conn, hash_value):
    cur = conn.execute("SELECT 1 FROM seen_leads WHERE hash = ?", (hash_value,))
    return cur.fetchone() is not None


def mark_seen(conn, hash_value):
    conn.execute(
        "INSERT OR IGNORE INTO seen_leads (hash, created_at) VALUES (?, ?)",
        (hash_value, datetime.now(timezone.utc).isoformat()),
    )
    conn.commit()


def _lead_row(lead):
    return [
        lead.get("date", ""),
        lead.get("source", ""),
        lead.get("group_name", ""),
        lead.get("author", ""),
        lead.get("text", ""),
        lead.get("link", ""),
        ", ".join(lead.get("matched_keywords", [])),
        lead.get("confidence", ""),
        lead.get("status", "новый"),
    ]


def append_lead_csv(lead, csv_path="leads.csv"):
    file_exists = os.path.isfile(csv_path)
    with open(csv_path, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(CSV_COLUMNS)
        writer.writerow(_lead_row(lead))


def append_lead_xlsx(lead, xlsx_path="leads.xlsx"):
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font
    except ImportError:
        logger.warning("openpyxl не установлен — пропускаю запись в XLSX (CSV создан)")
        return

    try:
        if os.path.isfile(xlsx_path):
            wb = load_workbook(xlsx_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"
            ws.append(CSV_COLUMNS)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.freeze_panes = "A2"

        ws.append(_lead_row(lead))

        for col_cells in ws.columns:
            length = max(
                (len(str(c.value)) if c.value is not None else 0) for c in col_cells
            )
            col_letter = col_cells[0].column_letter
            ws.column_dimensions[col_letter].width = min(max(length + 2, 12), 60)

        wb.save(xlsx_path)
    except Exception as exc:
        logger.warning("Не удалось обновить XLSX (%s): %s", xlsx_path, exc)


def save_lead(conn, lead, csv_path="leads.csv", xlsx_path="leads.xlsx"):
    """Возвращает True, если лид новый и был сохранён, False — если дубликат."""
    h = lead_hash(lead.get("author"), lead.get("text"))
    if is_duplicate(conn, h):
        return False
    append_lead_csv(lead, csv_path)
    append_lead_xlsx(lead, xlsx_path)
    mark_seen(conn, h)
    return True
